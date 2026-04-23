from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from io import BytesIO

def generate_firs_excel(data):

    wb = Workbook()
    ws = wb.active

    bold = Font(bold=True)
    red = Font(color="FF0000", bold=True)
    italic = Font(italic=True)
    center = Alignment(horizontal="center")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    try:
        logo = XLImage("assets/firs_logo.png")      

        # 🔥 Reduce size
        logo.width = 100   # adjust (try 80–150)
        logo.height = 50   # adjust (try 40–80)

        ws.add_image(logo, "C2")
        
    except:
        pass

    ws.merge_cells("A5:F5")
    ws["A5"] = "NOTE: FOR VAT STATUS COLUMN 0->VATABLE, 1-> ZERO RATED, 2->VAT EXEMPT"
    ws["A5"].font = red
    ws["A5"].alignment = center

    ws.merge_cells("A6:F6")
    ws["A6"] = "...It pays to pay your tax"
    ws["A6"].font = italic
    ws["A6"].alignment = center

    ws.merge_cells("A7:F7")
    ws["A7"] = "FIRS VAT SALE SCHEDULE"
    ws["A7"].font = Font(bold=True, size=16)
    ws["A7"].alignment = center

    headers = [
        "beneficiary_name","beneficiary_tin",
        "item","item_cost","item_description","vat_status"
    ]

    start = 8

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=i)
        cell.value = h
        cell.font = bold
        cell.border = border

    for r, row in enumerate(data, start=start+1):
        for c, k in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c)
            cell.value = row[k]
            cell.border = border

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer